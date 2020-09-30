# -*- coding: utf-8 -*-
"""
Created on Thu May 28 11:44:09 2020

@author: 15-db0011
"""
from ortools.linear_solver import pywraplp
import openpyxl


def main ():
    solver = pywraplp.Solver ( 'SolveAssignmentProblemMIP' , 
                           pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING )
    Deposito =[1,2]
    Cliente=[3,4,5,6,7]
    Nnodos=Deposito+Cliente
    #Nnodos=len(Nnodos)
    

    #Tabla de costos
    c=   [[5,   2,   4,   7,   3],
          [6,   3,   4,   5,   8]]
                                                   
    
    #Tabla de distancias 
    d=[[20,  10,  15,  10,  5],
       [15,  10,  5,   20,  30]]
    
    
    #Costo asociado al deposito 
    #o=[[50],
    #  [60]]
    
    
    Ovpn=[[20],
          [30]]
    
    
    #Capacidad  
    w=[210 , 300]
    
    #Demanda cliente
    dmd=[10, 5, 7, 10, 5]
    
    
    #capacidad del vehiculo en toneladas
    Q = 100
    #costo del vehiculo
    F =5
    #valor presente neto del costo del vehiculo
    Fvpn =20
    #cantidad de energia por distancia de un vehiculo sin carga
    h = 15.81
    #energia adicional por unidad de distancia y ton de carga de un vehiculo km por galon
    m = 12
    #emision total en kg de Co2 por galon
    E =  8.70645
    #Contador
    #cont=1
    

    x={}
    f={}
    a={}
    t={}
    y={}
    z={}
    
    for i in Nnodos:
          for j in Nnodos: #or for k in clientes
              x[ i , j ] = solver.BoolVar ( 'x [% i,% i]' % ( i , j ))
              a[ i , j ] = solver.BoolVar ( 'a [% i,% i]' % ( i , j ))
              t[ i , j ] = solver.NumVar  ( 0,solver.Infinity(),'t [% i,% i]' % ( i , j ))

    for i in Deposito:
        for j in Cliente:
            f[ i , j ] = solver.BoolVar  ( 'f [% i,% i]' % ( i , j ))  
                    
    for i in Deposito:
        y [ i ] = solver.BoolVar ( 'y [% i]' % ( i ))
        
    for j in Cliente:
        z[ j ] = solver.BoolVar ( 'z [% i]' % ( j ))
    #14    
    for j in Cliente:
        solver.Add(solver.Sum([x[i,j] for i in Nnodos])==1)
    
    #15
    for j in Cliente:
        solver.Add(solver.Sum([x[j,k] for k in Cliente])+ 
                  solver.Sum([a[i,j] for i in Deposito])== 
                  solver.Sum([x[i,j] for i in Nnodos]))
    
    #16
    for i in Deposito:
        solver.Add(solver.Sum([x[i, j] for j in Cliente])==
                            solver.Sum([a[i,j] for j in Cliente]))
                           
    #17
    for i in Nnodos:
        for j in Nnodos:
            solver.Add(x[i,j]+x[j,i]<=1) 
    
    #18
    for j in Cliente:
        solver.Add(solver.Sum([t[i,j] for i in Nnodos if i!=j])==
                  solver.Sum([t[j,k] for k in Nnodos if k!=j])+ dmd[j-len(Deposito)-1])
    
    #19
        solver.Add(solver.Sum([x[i,j] for i in Nnodos for j in Nnodos])== len(Nnodos))
    
    #20
    for j in Cliente:
        solver.Add(solver.Sum([f[i,j] for i in Deposito]) <= 1 ) 
        
    #21
    for i in Nnodos:
        for j in Nnodos:
            solver.Add(t[i,j] <= Q*x[i,j])
    
    #22
    for i in Deposito:
        solver.Add(solver.Sum([t[i,j] for j in Cliente])<=w[i-1]*y[i]) 
    
    #23
    for j in Cliente:
        solver.Add(solver.Sum([x[j,k] for k in Nnodos])==1-z[j])
    
    #24
    for i in Deposito:
        for j in Cliente:
            solver.Add(1+a[i,j]>=f[i,j]+z[j]) 
    
    #25
    for i in Deposito:
        for j in Cliente:
            for u in Cliente:
                solver.Add(x[j,u]+x[u,j]-1 <= f[i,j]-f[i,u])
    
    #26
    for i in Deposito:
        for j in Cliente:
            for u in Cliente:
                solver.Add(f[i,j]-f[i,u]<=1-x[j,u]-x[u,j]) 
    
    #27
    for i in Deposito:
        for j in Cliente:
            solver.Add(f[i,j]>=x[i,j]) 
            
    #28
    for i in Deposito:
        solver.Add(solver.Sum([y[i] for i in Deposito]) * sum([w[i-1] for i in Deposito]) >= sum([dmd[j-len(Deposito)-1] for j in Cliente])) 
    
    #29
    solver.Add(solver.Sum([x[i,j] for j in Cliente])<=w[i-1]/Q) 
    #30
    solver.Add(solver.Sum([x[i,j] for i in Deposito for j in Cliente])>=
    solver.Sum([dmd[j-len(Deposito)-1] for j in Cliente])/Q)
    
    solver.Minimize(solver.Sum([Ovpn[i-1][0] * y[i] for i in Deposito]) +
                   solver.Sum([Fvpn * a[i,j] for i in Deposito for j in Cliente])+ 
                   solver.Sum([c[i-1][j-len(Deposito)-1] * x[i,j] for i in Deposito for j in Cliente])+
                   solver.Sum([c[i-1][j-len(Deposito)-1] * a[i,j] for i in Deposito for j in Cliente])+
                   h*E*(solver.Sum([d[i-1][j-len(Deposito)-1] * x[i,j] for i in Deposito for j in Cliente])+
                        solver.Sum([d[i-1][j-len(Deposito)-1] * a[i,j] for i in Deposito for j in Cliente]))+
                   m*E*(solver.Sum([d[i-1][j-len(Deposito)-1] * t[i,j] for i in Deposito for j in Cliente])))
        
    status =solver.Solve()
    print(status)
    #0:  Optimo
    #1: Factible
    #2: Infactible
    #3: No acotado
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        assert solver.VerifySolution(1e-7, True)
        guardar(Deposito, Cliente, x, y , f, z, a, t, solver, Nnodos)
        if status == pywraplp.Solver.OPTIMAL:
            print('Optimal solution found')
        else:
            print('Feasible solution found')

    else:
        print('Problem not feasible or unbounded')
           
def guardar(Deposito, Cliente, x, y , f, z, a, t, solver, Nnodos):
    wb = openpyxl.load_workbook('Solucion.xlsx')
    wsLoc= wb['Loc']

    fila=1
    for i in Nnodos:
        for j in Nnodos:
            fila=fila+1
            wsLoc.cell(row = fila , column= 1).value = 'i'+ str(i)+'j'+ str(j)
            wsLoc.cell(row = fila , column= 2).value = x[i,j].solution_value()
            wsLoc.cell(row = fila , column= 3).value = a[i,j].solution_value()
            wsLoc.cell(row = fila , column= 4).value = t[i,j].solution_value()
            
    fila2=1        
    for i in Deposito:
        for j in Cliente:
            fila2=fila2+1
            wsLoc.cell(row = fila2 , column= 5).value = 'i'+ str(i)+'j'+ str(j) 
            wsLoc.cell(row = fila2 , column= 6).value = f[i,j].solution_value()
    
    for i in Deposito:
        wsLoc.cell(row = i+1 , column= 7).value = 'i'+ str(i)
        wsLoc.cell(row = i+1 , column= 8).value = y[i].solution_value()
           

       
    for j in Cliente:
        wsLoc.cell(row = j-len(Deposito)+1 , column= 9).value ='j'+ str(j)
        wsLoc.cell(row = j-len(Deposito)+1 , column= 10).value = z[j].solution_value()
    
    wsLoc.cell(row = 2 , column= 12).value = solver.Objective().Value()
    wb.save('Solucion.xlsx')
   
if __name__ == '__main__':
    main()
   
 

    
                
                
     
